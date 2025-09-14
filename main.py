import os
import barcode
import jdatetime
from barcode.writer import ImageWriter
from datetime import datetime

import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from tkinter import Toplevel

from openpyxl import Workbook, load_workbook

# --- تنظیمات اولیه پنجره اصلی ---
root = ttk.Window(themename="litera")
root.title("مدیریت نیروی انسانی")
root.geometry("1000x700")
root.resizable(False, False)
root.withdraw()  # مخفی‌کردن پنجره اصلی تا زمان ورود موفق

# --- متغیرهای ورودی ---
fname_var = ttk.StringVar()
lname_var = ttk.StringVar()
pedar_var = ttk.StringVar()
meli_var = ttk.StringVar()
shenasi_var = ttk.StringVar()
vaz_var = ttk.StringVar()

troz_var = ttk.StringVar()
tmah_var = ttk.StringVar()
tsal_var = ttk.StringVar()
oroz_var = ttk.StringVar()
omah_var = ttk.StringVar()
osal_var = ttk.StringVar()

delete_meli_var = ttk.StringVar()
search_meli_var = ttk.StringVar()
def update_clock():
    now = jdatetime.datetime.now()
    # قالب تاریخ شمسی + زمان دقیق
    time_str = now.strftime("%Y/%m/%d - %H:%M:%S")
    clock_label.config(text=time_str)
    clock_label.after(1000, update_clock)  # هر ۱۰۰۰ میلی ثانیه دوباره اجرا میشه

# ... در جایی از کد اصلی (بعد از ساخته شدن root) مثلا پایین‌تر از تعریف root:

clock_label = ttk.Label(root, font=("B Nazanin", 16, "bold"), foreground="blue")
clock_label.pack(side="top", pady=10)

update_clock()
# --- توابع اکسل ---
def save_to_excel(filename, data_dict, sheet_name="داده‌ها"):
    try:
        file_exists = os.path.exists(filename)
        if file_exists:
            wb = load_workbook(filename)
            ws = wb.active
            if ws.max_row == 0:
                ws.append(list(data_dict.keys()))
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            ws.append(list(data_dict.keys()))
        ws.append(list(data_dict.values()))
        wb.save(filename)
        return True
    except Exception as e:
        print("خطا در ذخیره فایل:", e)
        return False

# --- ذخیره و بارکد ---
def save_data_and_generate_barcode():
    status_label_add.config(text="⏳ لطفاً صبر کنید...", bootstyle="info")
    root.after(800, do_save_and_generate)

def do_save_and_generate():
    data_dict = {
        "نام": fname_var.get().strip(),
        "فامیلی": lname_var.get().strip(),
        "نام پدر": pedar_var.get().strip(),
        "کد ملی": meli_var.get().strip(),
        "شناسه بسیج": shenasi_var.get().strip(),
        "وضعیت عضویت": vaz_var.get().strip(),
        "روز تولد": troz_var.get().strip(),
        "ماه تولد": tmah_var.get().strip(),
        "سال تولد": tsal_var.get().strip(),
        "روز عضویت": oroz_var.get().strip(),
        "ماه عضویت": omah_var.get().strip(),
        "سال عضویت": osal_var.get().strip(),
    }

    empty = [k for k, v in data_dict.items() if v == ""]
    if empty:
        status_label_add.config(text=f"❌ لطفاً فیلد‌ها را کامل کنید: {', '.join(empty)}", bootstyle="danger")
        return

    ok = save_to_excel("member_data.xlsx", data_dict)
    if ok:
        status_label_add.config(text="✅ ذخیره با موفقیت انجام شد", bootstyle="success")
    else:
        status_label_add.config(text="❌ خطا در ذخیره اطلاعات", bootstyle="danger")

    code = data_dict["کد ملی"]
    if code:
        try:
            barcode_class = barcode.get_barcode_class('code128')
            my_barcode = barcode_class(code, writer=ImageWriter())
            my_barcode.save("barcode_image")
        except Exception as e:
            print("خطا در ساخت بارکد:", e)

# --- جستجو ---
def search_person_by_national_code():
    status_label_search.config(text="⏳ در حال جستجو...", bootstyle="info")
    root.after(800, do_search_person)

def do_search_person():
    nc = search_meli_var.get().strip()
    if nc == "":
        status_label_search.config(text="❌ لطفاً کد ملی را وارد کنید", bootstyle="danger")
        return

    try:
        wb = load_workbook("member_data.xlsx")
        ws = wb.active
    except FileNotFoundError:
        status_label_search.config(text="❌ فایل داده پیدا نشد", bootstyle="danger")
        return

    header = [cell.value for cell in ws[1]]
    found_row = None
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=4).value
        if val is not None and str(val).strip() == nc:
            found_row = r
            break

    if not found_row:
        status_label_search.config(text="❌ فردی با این کد ملی پیدا نشد", bootstyle="danger")
        return

    win = Toplevel(root)
    win.title("اطلاعات فرد")
    win.geometry("400x500")
    for i, col in enumerate(header, start=1):
        val = ws.cell(row=found_row, column=i).value
        ttk.Label(win, text=f"{col}:", font=("B Nazanin", 14, "bold")).grid(row=i, column=0, sticky="e", padx=10, pady=5)
        ttk.Label(win, text=f"{val}", font=("B Nazanin", 14)).grid(row=i, column=1, sticky="w", padx=10, pady=5)

    status_label_search.config(text="✅ فرد پیدا شد", bootstyle="success")

# --- حذف ---
def delete_person_by_national_code():
    status_label_del.config(text="⏳ لطفاً صبر کنید...", bootstyle="info")
    root.after(800, do_delete_person)

def do_delete_person():
    nc = delete_meli_var.get().strip()
    if nc == "":
        status_label_del.config(text="❌ لطفاً کد ملی را وارد کنید", bootstyle="danger")
        return

    try:
        wb = load_workbook("member_data.xlsx")
        ws = wb.active
    except FileNotFoundError:
        status_label_del.config(text="❌ فایل داده پیدا نشد", bootstyle="danger")
        return

    row_to_delete = None
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=4).value
        if val is not None and str(val).strip() == nc:
            row_to_delete = r
            break

    if row_to_delete:
        ws.delete_rows(row_to_delete, 1)
        wb.save("member_data.xlsx")
        status_label_del.config(text="✅ فرد با موفقیت حذف شد", bootstyle="success")
    else:
        status_label_del.config(text="❌ فردی با این کد ملی پیدا نشد", bootstyle="danger")

# --- طراحی صفحات ---
notebook = ttk.Notebook(root)
frame_add = ttk.Frame(notebook, padding=20)
frame_search = ttk.Frame(notebook, padding=20)
frame_delete = ttk.Frame(notebook, padding=20)
notebook.add(frame_add, text="اضافه کردن عضو")
notebook.add(frame_search, text="جستجو")
notebook.add(frame_delete, text="حذف عضو")
notebook.pack(fill="both", expand=True)

# --- فرم افزودن ---
labels = ["نام", "فامیلی", "نام پدر", "کد ملی",
          "شناسه بسیج", "وضعیت عضویت", "روز تولد", "ماه تولد",
          "سال تولد", "روز عضویت", "ماه عضویت", "سال عضویت"]
vars_ = [fname_var, lname_var, pedar_var, meli_var,
         shenasi_var, vaz_var, troz_var, tmah_var,
         tsal_var, oroz_var, omah_var, osal_var]

for i, (lbl, var) in enumerate(zip(labels, vars_)):
    row = i // 2
    col = (i % 2) * 2
    ttk.Label(frame_add, text=lbl + ":", font=("B Nazanin", 14)).grid(row=row, column=col, sticky="e", padx=10, pady=5)
    ttk.Entry(frame_add, textvariable=var, font=("B Nazanin", 14)).grid(row=row, column=col+1, sticky="we", padx=10, pady=5)

status_label_add = ttk.Label(frame_add, text="", font=("B Nazanin", 14))
status_label_add.grid(row=7, column=0, columnspan=4, pady=(20,0))

ttk.Button(frame_add, text="ذخیره و ساخت بارکد", bootstyle=SUCCESS, command=save_data_and_generate_barcode).grid(row=8, column=0, columnspan=4, pady=20)

# --- فرم جستجو ---
ttk.Label(frame_search, text="کد ملی برای جستجو:", font=("B Nazanin", 14)).grid(row=0, column=0, padx=10, pady=10, sticky="e")
ttk.Entry(frame_search, textvariable=search_meli_var, font=("B Nazanin", 14)).grid(row=0, column=1, padx=10, pady=10, sticky="we")
status_label_search = ttk.Label(frame_search, text="", font=("B Nazanin", 14))
status_label_search.grid(row=1, column=0, columnspan=2, pady=(5,0))
ttk.Button(frame_search, text="جستجو", bootstyle=INFO, command=search_person_by_national_code).grid(row=2, column=0, columnspan=2, pady=20)

# --- فرم حذف ---
ttk.Label(frame_delete, text="کد ملی برای حذف:", font=("B Nazanin", 14)).grid(row=0, column=0, padx=10, pady=10, sticky="e")
ttk.Entry(frame_delete, textvariable=delete_meli_var, font=("B Nazanin", 14)).grid(row=0, column=1, padx=10, pady=10, sticky="we")
status_label_del = ttk.Label(frame_delete, text="", font=("B Nazanin", 14))
status_label_del.grid(row=1, column=0, columnspan=2, pady=(5,0))
ttk.Button(frame_delete, text="حذف عضو", bootstyle=DANGER, command=delete_person_by_national_code).grid(row=2, column=0, columnspan=2, pady=20)

# --- پنجره ورود ---
def check_password():
    password = password_var.get()
    if password == "936330":  # 🔐 رمز عبور
        login_win.destroy()
        root.deiconify()  # نمایش برنامه اصلی
    else:
        login_status.config(text="❌ رمز اشتباه است", bootstyle="danger")

# ساخت پنجره لاگین
login_win = Toplevel()
login_win.title("ورود")
login_win.geometry("500x400")
login_win.resizable(False, False)
login_win.grab_set()

password_var = ttk.StringVar()

ttk.Label(login_win, text="رمز عبور:", font=("B Nazanin", 14)).pack(pady=20)
ttk.Entry(login_win, textvariable=password_var, font=("B Nazanin", 14), show="*").pack(pady=5)

login_status = ttk.Label(login_win, text="", font=("B Nazanin", 12))
login_status.pack(pady=5)

ttk.Button(login_win, text="ورود", bootstyle=PRIMARY, command=check_password).pack(pady=10)

login_win.protocol("WM_DELETE_WINDOW", root.destroy)  # بستن کل برنامه اگر پنجره ورود بسته شد

# --- اجرای برنامه ---
root.mainloop()
